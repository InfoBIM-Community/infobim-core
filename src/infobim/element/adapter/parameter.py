from __future__ import annotations

import json
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse
from urllib.request import url2pathname

from ontobdc.storage.adapter.bootstrap import (
    get_container_storage_file_path,
    get_ontobdc_directory,
)
from openpyxl import load_workbook
from rdflib import Graph, Literal, Namespace, URIRef
from rdflib.namespace import PROV, RDF

OBDC = Namespace("http://ontobdc.org/ontology/domain/ns.ttl#")


@dataclass(frozen=True)
class ElementParameterMutation:
    global_id: str
    parameter_uri: str
    old_value: Any
    new_value: Any
    dataset_path: str
    workbook_path: str
    worksheet: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "global_id": self.global_id,
            "parameter_uri": self.parameter_uri,
            "old_value": self.old_value,
            "new_value": self.new_value,
            "dataset_path": self.dataset_path,
            "workbook_path": self.workbook_path,
            "worksheet": self.worksheet,
        }


@dataclass(frozen=True)
class _FieldContract:
    uri: str
    identifier: str
    datatype: str
    identifier_field: bool


@dataclass(frozen=True)
class _CellBinding:
    global_id: str
    parameter: _FieldContract
    value: Any
    dataset_path: Path
    workbook_path: Path
    worksheet: str
    row: int
    column: int


class ElementParameterRepository:
    """Mutate generic entity-instance fields declared by dataset facades.

    Dataset discovery follows the Project's semantic container metadata. It
    never scans the filesystem for arbitrary workbooks and never assumes an
    IFC entity type: an Element is an entity instance identified by GlobalId.
    """

    def __init__(self, project_path: str) -> None:
        self._project_path = Path(project_path).expanduser().resolve()
        if not self._project_path.is_dir():
            raise ValueError(f"InfoBIM Project path not found: {project_path}")

    def set(
        self,
        global_id: str,
        parameter_uri: str,
        value: str,
        *,
        all_occurrences: bool = False,
    ) -> list[ElementParameterMutation]:
        bindings = self._parameter_bindings(global_id, parameter_uri)
        if not all_occurrences and len(bindings) != 1:
            raise ValueError(
                f"Parameter URI '{parameter_uri}' resolves {len(bindings)} times "
                f"for element '{global_id}'; use --all to update every occurrence."
            )
        selected = bindings if all_occurrences else bindings[:1]
        coerced = [
            self._coerce(value, binding.parameter.datatype) for binding in selected
        ]
        return self._write(selected, coerced)

    def unset(
        self,
        global_id: str,
        parameter_uri: str,
    ) -> ElementParameterMutation:
        bindings = self._parameter_bindings(global_id, parameter_uri)
        if len(bindings) != 1:
            raise ValueError(
                f"Parameter URI '{parameter_uri}' resolves {len(bindings)} times "
                f"for element '{global_id}'."
            )
        return self._write(bindings, [None])[0]

    def unset_all(self, global_id: str) -> list[ElementParameterMutation]:
        bindings = [
            binding
            for binding in self._bindings_for_global_id(global_id)
            if not binding.parameter.identifier_field
        ]
        if not bindings:
            raise ValueError(
                f"Element with GlobalId '{global_id}' has no writable parameters."
            )
        return self._write(bindings, [None] * len(bindings))

    def _parameter_bindings(
        self,
        global_id: str,
        parameter_uri: str,
    ) -> list[_CellBinding]:
        target = str(parameter_uri or "").strip()
        if not target:
            raise ValueError("--parameter requires a URI.")
        bindings = [
            binding
            for binding in self._bindings_for_global_id(global_id)
            if binding.parameter.uri == target
        ]
        if not bindings:
            raise ValueError(
                f"Parameter URI '{target}' is not declared for element '{global_id}'."
            )
        if any(binding.parameter.identifier_field for binding in bindings):
            raise ValueError(
                f"Parameter '{target}' identifies the element and cannot be changed."
            )
        return bindings

    def _bindings_for_global_id(self, global_id: str) -> list[_CellBinding]:
        normalized_id = str(global_id or "").strip()
        if not normalized_id:
            raise ValueError("--global-id is required.")

        bindings: list[_CellBinding] = []
        for dataset_path in self._entity_dataset_paths():
            contracts = self._field_contracts(dataset_path)
            if not contracts:
                continue
            datapackage_path = get_ontobdc_directory(dataset_path) / "datapackage.json"
            if not datapackage_path.is_file():
                continue
            descriptor = self._read_json(datapackage_path)
            for resource in descriptor.get("resources", []):
                if not isinstance(resource, dict):
                    continue
                if str(resource.get("format") or "").lower() != "xlsx":
                    continue
                workbook_path = self._resource_path(datapackage_path, resource)
                worksheet = str(
                    ((resource.get("dialect") or {}).get("excel") or {}).get("sheet")
                    or resource.get("name")
                    or ""
                ).strip()
                bindings.extend(
                    self._workbook_bindings(
                        dataset_path=dataset_path,
                        workbook_path=workbook_path,
                        worksheet=worksheet,
                        global_id=normalized_id,
                        contracts=contracts,
                    )
                )
        if not bindings:
            raise ValueError(
                f"Element with GlobalId '{normalized_id}' was not found in the Project."
            )
        return sorted(
            bindings,
            key=lambda item: (
                item.parameter.uri,
                str(item.dataset_path),
                str(item.workbook_path),
                item.worksheet,
            ),
        )

    def _entity_dataset_paths(self) -> Iterable[Path]:
        metadata_path = get_container_storage_file_path(self._project_path)
        if not metadata_path.is_file():
            raise ValueError(f"Project container metadata not found: {metadata_path}")
        graph = Graph().parse(metadata_path, format="turtle")
        containers = list(graph.subjects(RDF.type, OBDC.DataContainer))
        if len(containers) != 1:
            raise ValueError(
                f"Project metadata must declare exactly one DataContainer: {metadata_path}"
            )
        for dataset in graph.objects(containers[0], OBDC.hasEntityDataset):
            location = graph.value(dataset, PROV.atLocation)
            if location is None:
                continue
            yield self._location_path(str(location))

    def _location_path(self, location: str) -> Path:
        parsed = urlparse(location)
        if parsed.scheme == "file":
            return Path(url2pathname(unquote(parsed.path))).expanduser().resolve()
        path = Path(location).expanduser()
        if not path.is_absolute():
            path = self._project_path / path
        return path.resolve()

    def _field_contracts(self, dataset_path: Path) -> dict[str, _FieldContract]:
        facade_path = get_ontobdc_directory(dataset_path) / "linkset" / "facade.ttl"
        if not facade_path.is_file():
            return {}
        graph = Graph().parse(facade_path, format="turtle")
        contracts: dict[str, _FieldContract] = {}
        for subject in {
            value for value in graph.subjects() if isinstance(value, URIRef)
        }:
            identifier = self._first_literal(graph, subject, "identifier")
            mapped = self._first_uri(graph, subject, "mapsToProperty")
            if identifier is None or mapped is None:
                continue
            identifier_text = str(identifier).strip()
            datatype = self._first_uri(graph, subject, "fieldDatatype")
            is_identifier = self._first_literal(graph, subject, "isIdentifierField")
            contracts[self._normalize(identifier_text)] = _FieldContract(
                uri=str(mapped),
                identifier=identifier_text,
                datatype=self._local_name(datatype) if datatype else "string",
                identifier_field=str(is_identifier or "false").lower() == "true",
            )
        return contracts

    def _workbook_bindings(
        self,
        *,
        dataset_path: Path,
        workbook_path: Path,
        worksheet: str,
        global_id: str,
        contracts: dict[str, _FieldContract],
    ) -> list[_CellBinding]:
        if not workbook_path.is_file():
            return []
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        try:
            if worksheet not in workbook.sheetnames:
                return []
            sheet = workbook[worksheet]
            headers = [str(cell.value or "").strip() for cell in sheet[1]]
            id_column = next(
                (
                    index
                    for index, header in enumerate(headers, start=1)
                    if self._normalize(header) == "globalid"
                ),
                None,
            )
            if id_column is None:
                return []
            rows = [
                row
                for row in range(2, sheet.max_row + 1)
                if str(sheet.cell(row=row, column=id_column).value or "").strip()
                == global_id
            ]
            result: list[_CellBinding] = []
            for row in rows:
                for column, header in enumerate(headers, start=1):
                    contract = contracts.get(self._normalize(header))
                    if contract is None:
                        continue
                    result.append(
                        _CellBinding(
                            global_id=global_id,
                            parameter=contract,
                            value=sheet.cell(row=row, column=column).value,
                            dataset_path=dataset_path,
                            workbook_path=workbook_path,
                            worksheet=worksheet,
                            row=row,
                            column=column,
                        )
                    )
            return result
        finally:
            workbook.close()

    def _write(
        self,
        bindings: list[_CellBinding],
        values: list[Any],
    ) -> list[ElementParameterMutation]:
        by_workbook: dict[Path, list[tuple[_CellBinding, Any]]] = {}
        for binding, value in zip(bindings, values):
            by_workbook.setdefault(binding.workbook_path, []).append((binding, value))

        mutations: list[ElementParameterMutation] = []
        for workbook_path, changes in by_workbook.items():
            workbook = load_workbook(workbook_path)
            try:
                for binding, value in changes:
                    workbook[binding.worksheet].cell(
                        row=binding.row,
                        column=binding.column,
                    ).value = value
                    mutations.append(
                        ElementParameterMutation(
                            global_id=binding.global_id,
                            parameter_uri=binding.parameter.uri,
                            old_value=binding.value,
                            new_value=value,
                            dataset_path=str(binding.dataset_path),
                            workbook_path=str(workbook_path),
                            worksheet=binding.worksheet,
                        )
                    )
                workbook.save(workbook_path)
            finally:
                workbook.close()
        return mutations

    @staticmethod
    def _resource_path(datapackage_path: Path, resource: dict[str, Any]) -> Path:
        raw_path = str(resource.get("path") or "").strip()
        parsed = urlparse(raw_path)
        if parsed.scheme == "file":
            return Path(url2pathname(unquote(parsed.path))).expanduser().resolve()
        path = Path(raw_path).expanduser()
        if not path.is_absolute():
            path = datapackage_path.parent / path
        return path.resolve()

    @staticmethod
    def _coerce(value: str, datatype: str) -> Any:
        normalized = str(datatype or "string").lower()
        raw = str(value)
        if normalized in {"boolean", "bool"}:
            lowered = raw.strip().lower()
            if lowered in {"true", "1", "yes"}:
                return True
            if lowered in {"false", "0", "no"}:
                return False
            raise ValueError(f"Invalid boolean value: {value}")
        if normalized in {"integer", "int", "positiveinteger"}:
            return int(raw)
        if normalized in {"decimal", "double", "float", "number"}:
            return float(raw)
        return raw

    @staticmethod
    def _read_json(path: Path) -> dict[str, Any]:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise ValueError(f"Invalid datapackage: {path}") from error
        if not isinstance(payload, dict):
            raise TypeError(f"Invalid datapackage: {path}")
        return payload

    @classmethod
    def _first_literal(
        cls,
        graph: Graph,
        subject: URIRef,
        local_name: str,
    ) -> Literal | None:
        return next(
            (
                value
                for _, predicate, value in graph.triples((subject, None, None))
                if cls._local_name(predicate) == local_name
                and isinstance(value, Literal)
            ),
            None,
        )

    @classmethod
    def _first_uri(
        cls,
        graph: Graph,
        subject: URIRef,
        local_name: str,
    ) -> URIRef | None:
        return next(
            (
                value
                for _, predicate, value in graph.triples((subject, None, None))
                if cls._local_name(predicate) == local_name
                and isinstance(value, URIRef)
            ),
            None,
        )

    @staticmethod
    def _normalize(value: object) -> str:
        return "".join(
            character for character in str(value).lower() if character.isalnum()
        )

    @staticmethod
    def _local_name(value: object) -> str:
        raw = str(value or "")
        if "#" in raw:
            return raw.rsplit("#", 1)[-1]
        return raw.rstrip("/").rsplit("/", 1)[-1]
