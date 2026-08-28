"""Export an IfcWorkSchedule workbook (IfcTask + IfcTaskTime sheets) to a
paginated PDF: a task table (WBS, name, start, finish, duration, status)
next to a matching Gantt chart, one page per chunk of tasks so a long
schedule stays readable.

This adapter only translates workbook data into a document. Container
resolution belongs to ``ScheduleWorkbookAdapter`` and orchestration belongs
to the 4D capability and command layers.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from matplotlib.figure import Figure

TASK_SHEET: str = "IfcTask"
TASK_TIME_SHEET: str = "IfcTaskTime"
TASKS_PER_PAGE: int = 25

_ISO_DURATION_DAYS: re.Pattern[str] = re.compile(r"^P(?:(\d+)D)?")


@dataclass
class ScheduleTask:
    identification: str
    name: str
    start: Optional[dt.datetime]
    finish: Optional[dt.datetime]
    duration_days: Optional[int]
    completion: Optional[float]
    status: str
    is_milestone: bool


class ScheduleGanttPdfExporter:
    """Read one IfcWorkSchedule workbook and render it to a PDF."""

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path: Path = Path(workbook_path).expanduser().resolve()
        if not self.workbook_path.is_file():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

    def tasks(self) -> List[ScheduleTask]:
        workbook: Workbook = load_workbook(
            filename=str(self.workbook_path), data_only=True
        )
        try:
            task_rows: List[Dict[str, Any]] = self._records(workbook, TASK_SHEET)
            time_rows_by_id: Dict[str, Dict[str, Any]] = {
                str(row.get("GlobalId") or ""): row
                for row in self._records(workbook, TASK_TIME_SHEET)
            }
        finally:
            workbook.close()

        tasks: List[ScheduleTask] = []
        for task_row in task_rows:
            task_time_id: str = str(
                task_row.get("TaskTime")
                or task_row.get("TaskTimeGlobalId")
                or ""
            )
            time_row: Dict[str, Any] = time_rows_by_id.get(task_time_id, {})
            start: Optional[dt.datetime] = self._parse_datetime(
                time_row.get("ScheduleStart")
            )
            finish: Optional[dt.datetime] = self._parse_datetime(
                time_row.get("ScheduleFinish")
            )
            tasks.append(
                ScheduleTask(
                    identification=str(task_row.get("Identification") or ""),
                    name=str(task_row.get("Name") or ""),
                    start=start,
                    finish=finish,
                    duration_days=self._duration_days(
                        start, finish, time_row.get("ScheduleDuration")
                    ),
                    completion=self._parse_completion(
                        time_row.get("PercentComplete")
                        if time_row.get("PercentComplete") is not None
                        else time_row.get("Completion")
                    ),
                    status=str(task_row.get("Status") or ""),
                    is_milestone=self._truthy(task_row.get("IsMilestone")),
                )
            )
        return sorted(tasks, key=self._task_sort_key)

    @staticmethod
    def _task_sort_key(task: ScheduleTask) -> List[Any]:
        parts: List[Any] = []
        for part in task.identification.split("."):
            stripped: str = part.strip()
            parts.append(
                (0, int(stripped)) if stripped.isdigit() else (1, stripped)
            )
        return parts

    def export(self, out_path: Path) -> Path:
        from matplotlib.backends.backend_pdf import PdfPages

        tasks: List[ScheduleTask] = self.tasks()
        out_path = Path(out_path).expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)

        with PdfPages(str(out_path)) as pdf:
            if not tasks:
                pdf.savefig(self._empty_page())
                return out_path

            page_count: int = (len(tasks) + TASKS_PER_PAGE - 1) // TASKS_PER_PAGE
            page_index: int
            for page_index in range(page_count):
                chunk: List[ScheduleTask] = tasks[
                    page_index * TASKS_PER_PAGE : (page_index + 1) * TASKS_PER_PAGE
                ]
                figure: Figure = self._page(
                    chunk,
                    page_number=page_index + 1,
                    page_count=page_count,
                    date_bounds=self._date_bounds(tasks),
                )
                pdf.savefig(figure)

        return out_path

    # ------------------------------------------------------------ reading

    @staticmethod
    def _records(workbook: Workbook, sheet_name: str) -> List[Dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet: Worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header: List[str] = [
                str(cell).strip() if cell is not None else "" for cell in next(rows)
            ]
        except StopIteration:
            return []

        records: List[Dict[str, Any]] = []
        for raw in rows:
            record: Dict[str, Any] = {
                column: value
                for column, value in zip(header, raw)
                if column and value is not None
            }
            if record:
                records.append(record)
        return records

    @staticmethod
    def _parse_datetime(value: Any) -> Optional[dt.datetime]:
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time())
        text: str = str(value).strip()
        if not text:
            return None
        try:
            return dt.datetime.fromisoformat(text)
        except ValueError:
            return None

    @classmethod
    def _duration_days(
        cls,
        start: Optional[dt.datetime],
        finish: Optional[dt.datetime],
        iso_duration: Any,
    ) -> Optional[int]:
        if start is not None and finish is not None:
            return max(0, (finish - start).days)
        match: Optional[re.Match[str]] = _ISO_DURATION_DAYS.match(
            str(iso_duration or "").strip()
        )
        if match and match.group(1):
            return int(match.group(1))
        return None

    @staticmethod
    def _parse_completion(value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            number: float = float(value)
        except (TypeError, ValueError):
            return None
        return number * 100 if 0 <= number <= 1 else number

    @staticmethod
    def _truthy(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        return str(value or "").strip().lower() in {"1", "true", "yes", "y"}

    # -------------------------------------------------------------- render

    @staticmethod
    def _date_bounds(tasks: List[ScheduleTask]):
        dates: List[dt.datetime] = [
            d for t in tasks for d in (t.start, t.finish) if d is not None
        ]
        if not dates:
            today = dt.datetime.combine(dt.date.today(), dt.time())
            return today, today + dt.timedelta(days=30)
        return min(dates), max(dates)

    def _page(
        self,
        chunk: List[ScheduleTask],
        *,
        page_number: int,
        page_count: int,
        date_bounds,
    ) -> Figure:
        from matplotlib.figure import Figure

        figure: Figure = Figure(figsize=(11.7, 8.3))  # A4 landscape
        figure.suptitle(
            f"{self.workbook_path.name} -- IfcWorkSchedule "
            f"(page {page_number}/{page_count})",
            fontsize=11,
            fontweight="bold",
        )
        table_axes = figure.add_axes((0.03, 0.06, 0.44, 0.85))
        gantt_axes = figure.add_axes((0.50, 0.06, 0.47, 0.85))

        self._draw_table(table_axes, chunk)
        self._draw_gantt(gantt_axes, chunk, date_bounds)
        return figure

    @staticmethod
    def _draw_table(axes, chunk: List[ScheduleTask]) -> None:
        axes.axis("off")
        columns = ["WBS", "Task", "Start", "Finish", "Days", "%"]
        rows = [
            [
                task.identification,
                (task.name[:38] + "...") if len(task.name) > 38 else task.name,
                task.start.strftime("%Y-%m-%d") if task.start else "-",
                task.finish.strftime("%Y-%m-%d") if task.finish else "-",
                str(task.duration_days) if task.duration_days is not None else "-",
                f"{task.completion:.0f}" if task.completion is not None else "-",
            ]
            for task in chunk
        ]
        table = axes.table(
            cellText=rows,
            colLabels=columns,
            colWidths=[0.14, 0.44, 0.15, 0.15, 0.08, 0.08],
            cellLoc="left",
            loc="upper left",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(7)
        table.scale(1, 1.35)
        for (row, _col), cell in table.get_celld().items():
            cell.set_edgecolor("#dddddd")
            if row == 0:
                cell.set_facecolor("#17365d")
                cell.set_text_props(color="white", fontweight="bold")

    @staticmethod
    def _draw_gantt(axes, chunk: List[ScheduleTask], date_bounds) -> None:
        import matplotlib.dates as mdates

        min_date, max_date = date_bounds
        row_count: int = len(chunk)
        for index, task in enumerate(chunk):
            y: int = row_count - 1 - index
            if task.start is None or task.finish is None:
                axes.text(
                    min_date, y, "no schedule", fontsize=6, color="#999999",
                    va="center",
                )
                continue
            if task.is_milestone or task.start == task.finish:
                axes.plot(
                    [task.start], [y], marker="D", color="#17365d", markersize=6,
                )
                continue

            axes.barh(
                y, (task.finish - task.start).days, left=task.start,
                height=0.55, color="#38bdf8", edgecolor="#17365d", linewidth=0.5,
            )
            if task.completion:
                progress_days = (task.finish - task.start).days * (
                    task.completion / 100
                )
                axes.barh(
                    y, progress_days, left=task.start, height=0.55,
                    color="#0f6fb0",
                )

        today = dt.datetime.combine(dt.date.today(), dt.time())
        if min_date <= today <= max_date:
            axes.axvline(today, color="#e53e3e", linewidth=1, linestyle="--")

        axes.set_ylim(-0.5, row_count - 0.5)
        axes.set_yticks([])
        span = max((max_date - min_date).days, 1)
        axes.set_xlim(
            min_date - dt.timedelta(days=max(1, span * 0.02)),
            max_date + dt.timedelta(days=max(1, span * 0.02)),
        )
        axes.xaxis.set_major_locator(mdates.AutoDateLocator())
        axes.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
        axes.tick_params(axis="x", labelrotation=45, labelsize=6)
        axes.grid(axis="x", color="#eeeeee")
        for spine in ("top", "right", "left"):
            axes.spines[spine].set_visible(False)

    @staticmethod
    def _empty_page() -> Figure:
        from matplotlib.figure import Figure

        figure = Figure(figsize=(11.7, 8.3))
        axes = figure.add_axes((0.1, 0.1, 0.8, 0.8))
        axes.axis("off")
        axes.text(0.5, 0.5, "No tasks found in this workbook.", ha="center")
        return figure
