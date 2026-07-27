import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from infobim.project.adapter.repository import ElementImportStepRepository
from infobim.project.domain.machine.import_state import ElementImportProcessState
from ontobdc.cli.domain.port.context import CliContextPort
from ontobdc.context.adapter.repository import LocalContextFileResource
from ontobdc.shared.adapter.capability import TransformationCapability
from ontobdc.shared.domain.model.capability import CapabilityMetadata


class GeneratedCapability(TransformationCapability):
    METADATA = CapabilityMetadata(
        id="org.infobim.project.plugin.capability.transformation.target.generated",
        version="1.0.0",
        name="Project element import transformation to Generated",
        description="Generate an Excel workbook from the instantiated IfcTask payload.",
        author=["TRAE"],
        tags=["project", "import", "generated", "xlsx", "ifctask"],
        supported_languages=["en", "pt-br"],
    )

    def label(self, lang: str = "en") -> str:
        return "Project element import transformation to Generated"

    def description(self, lang: str = "en") -> str:
        return "Generate an Excel workbook from the instantiated IfcTask payload."

    def execute(self, context: CliContextPort) -> Dict[str, Any]:
        try:
            from frictionless import Package, Resource, Schema, formats, validate
            from openpyxl import Workbook
            from openpyxl.cell import Cell
            from openpyxl.styles import Font
        except ModuleNotFoundError as exc:
            raise ValueError(
                "The 'openpyxl' and 'frictionless' packages are required to generate the workbook and datapackage."
            ) from exc

        step_repository: Optional[ElementImportStepRepository] = context.get_parameter_value("element_import_step_repository")
        if not isinstance(step_repository, ElementImportStepRepository):
            step_repository = ElementImportStepRepository(
                container_path=str(context.get_parameter_value("container_path")),
                source_path=str(context.get_parameter_value("import_path")),
                element_name=str(context.get_parameter_value("element_name")),
            )
            context.set_parameter_value("element_import_step_repository", step_repository)

        instantiated_resource: LocalContextFileResource = step_repository.reload(ElementImportProcessState.INSTANTIATED)
        instantiated_payload: Dict[str, Any] = json.loads(str(instantiated_resource.content))
        ifcjson_instances: List[Dict[str, Any]] = list(instantiated_payload.get("ifcjson_instances", []))
        if not ifcjson_instances:
            raise ValueError("The instantiated artifact does not expose any IfcTask payload to generate the workbook.")

        output_dir: Path = Path(str(context.get_parameter_value("container_path"))).expanduser().resolve() / "payload" / "document"
        output_dir.mkdir(parents=True, exist_ok=True)
        ontobdc_dir: Path = Path(str(context.get_parameter_value("container_path"))).expanduser().resolve() / ".__ontobdc__"
        workbook_path: Path = self._resolve_writable_workbook_path(
            output_dir=output_dir,
            workbook_name=self._build_workbook_name(step_repository),
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "IfcTask"

        headers: List[str] = [
            "GlobalId",
            "Identification",
            "Name",
            "IsMilestone",
            "TaskTime",
        ]
        self._append_literal_row(worksheet=worksheet, row_values=headers)
        for header_cell in worksheet[1]:
            header_cell.font = Font(bold=True)

        generated_row_count: int = 0
        for instance_payload in ifcjson_instances:
            task_time_payload: Dict[str, Any] = dict(instance_payload.get("TaskTime", {}))
            row_values: List[Any] = [
                instance_payload.get("GlobalId"),
                instance_payload.get("Identification"),
                instance_payload.get("Name"),
                instance_payload.get("IsMilestone"),
                self._serialize_task_time(task_time_payload),
            ]
            self._append_literal_row(
                worksheet=worksheet,
                row_values=row_values,
            )
            generated_row_count += 1

        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            column_letter: str = column_cells[0].column_letter
            max_length: int = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 64)

        workbook.save(workbook_path)
        datapackage_path: Path = ontobdc_dir / "datapackage.json"
        workbook_relative_path: str = Path("..", "payload", "document", workbook_path.name).as_posix()
        package = Package(
            name=self._build_datapackage_name(step_repository),
            basepath=str(ontobdc_dir),
            resources=[
                Resource(
                    name="ifc_task",
                    path=workbook_relative_path,
                    schema=Schema(
                        {
                            "fields": [
                                {"name": "GlobalId", "type": "string"},
                                {"name": "Identification", "type": "string"},
                                {"name": "Name", "type": "string"},
                                {"name": "IsMilestone", "type": "boolean"},
                                {"name": "TaskTime", "type": "string"},
                            ],
                            "primaryKey": ["GlobalId"],
                        }
                    ),
                    control=formats.ExcelControl(sheet=worksheet.title),
                )
            ],
        )
        package.to_json(str(datapackage_path))
        validation_report = validate(str(datapackage_path))
        validation_payload: Dict[str, Any] = validation_report.to_descriptor()
        if not validation_report.valid:
            raise ValueError(
                f"Frictionless validation failed for datapackage: {json.dumps(validation_payload, ensure_ascii=False)}"
            )

        generated_payload: Dict[str, Any] = {
            "source_state": ElementImportProcessState.INSTANTIATED.value,
            "source_artifact": str(instantiated_resource.path),
            "datapackage_path": str(datapackage_path),
            "workbook_path": str(workbook_path),
            "workbook_format": "xlsx",
            "worksheet_name": worksheet.title,
            "validation": validation_payload,
            "summary": {
                "generated_row_count": generated_row_count,
                "ifc_task_count": len(ifcjson_instances),
            },
        }
        generated_artifact_path: Path = step_repository.write_text_file(
            state=ElementImportProcessState.GENERATED,
            content=json.dumps(generated_payload, ensure_ascii=False, indent=2, sort_keys=True),
            file_type="json",
        )
        context.set_parameter_value("resource", LocalContextFileResource(generated_artifact_path))
        context.set_parameter_value("generated_workbook_path", str(workbook_path))
        return {
            "resulting_state": ElementImportProcessState.GENERATED,
            "path": str(generated_artifact_path),
            "datapackage_path": str(datapackage_path),
            "workbook_path": str(workbook_path),
            "generated_row_count": generated_row_count,
        }

    def _build_workbook_name(self, step_repository: ElementImportStepRepository) -> str:
        source_stem: str = re.sub(r"[^0-9A-Za-z_-]+", "_", step_repository.source_path.stem).strip("_")
        element_name: str = re.sub(r"[^0-9A-Za-z_-]+", "_", step_repository.element_name).strip("_")
        if not source_stem:
            source_stem = "document"
        if not element_name:
            element_name = "element"
        return f"{source_stem}__{element_name}__{step_repository.source_hash[:8]}.xlsx"

    def _build_datapackage_name(self, step_repository: ElementImportStepRepository) -> str:
        source_stem: str = re.sub(r"[^0-9A-Za-z_-]+", "_", step_repository.source_path.stem).strip("_").lower()
        element_name: str = re.sub(r"[^0-9A-Za-z_-]+", "_", step_repository.element_name).strip("_").lower()
        if not source_stem:
            source_stem = "document"
        if not element_name:
            element_name = "element"
        return f"{source_stem}_{element_name}"

    def _resolve_writable_workbook_path(self, output_dir: Path, workbook_name: str) -> Path:
        base_path: Path = output_dir / workbook_name
        if not base_path.exists():
            return base_path

        try:
            with base_path.open("ab"):
                return base_path
        except PermissionError:
            pass

        stem: str = base_path.stem
        suffix: str = base_path.suffix
        for index in range(1, 1000):
            candidate_path: Path = output_dir / f"{stem}__{index}{suffix}"
            if not candidate_path.exists():
                return candidate_path
            try:
                with candidate_path.open("ab"):
                    return candidate_path
            except PermissionError:
                continue

        raise PermissionError(f"Could not resolve a writable workbook path under: {output_dir}")

    def _append_literal_row(self, worksheet: Any, row_values: List[Any]) -> None:
        next_row_index: int = worksheet.max_row + 1
        if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
            next_row_index = 1
        for column_index, raw_value in enumerate(row_values, start=1):
            cell: Cell = worksheet.cell(row=next_row_index, column=column_index)
            self._assign_cell_literal_value(cell=cell, raw_value=raw_value)

    def _assign_cell_literal_value(self, cell: Any, raw_value: Any) -> None:
        if raw_value is None:
            return
        if isinstance(raw_value, bool):
            cell.value = raw_value
            return
        if isinstance(raw_value, (int, float)):
            cell.value = raw_value
            return

        cell.value = str(raw_value)
        cell.data_type = "s"

    def _serialize_task_time(self, task_time_payload: Dict[str, Any]) -> str:
        if not task_time_payload:
            return ""

        serialized_payload: Dict[str, Any] = {
            key: value
            for key, value in task_time_payload.items()
            if key != "type"
        }
        if not serialized_payload:
            return ""

        return json.dumps(serialized_payload, ensure_ascii=False, sort_keys=True)
