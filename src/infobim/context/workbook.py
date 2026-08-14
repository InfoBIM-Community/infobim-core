

import json
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from infobim.context.ontology import EntityContextContract, EntityFacadeContract
from infobim.project.plugin.capability.transformation.ifc_project_ready import (
    IfcProjectReadyCapability,
)


@dataclass(frozen=True)
class EntityContextWorkbookArtifact:
    workbook_path: Path
    datapackage_path: Path
    worksheets: tuple[str, ...]
    root_global_id: str


class EntityContextWorkbookAdapter:
    """Materialize a semantic entity context as one physical XLSX option."""

    def generate(
        self,
        *,
        dataset_path: Path,
        contract: EntityContextContract,
        instance_name: str,
    ) -> EntityContextWorkbookArtifact:
        output_dir = dataset_path.resolve() / "payload" / "document"
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = output_dir / (
            f"{self._snake_case(contract.root.entity_name)}.xlsx"
        )
        datapackage_path = dataset_path / ".__ontobdc__" / "datapackage.json"

        workbook = Workbook()
        workbook.remove(workbook.active)
        root_global_id = IfcProjectReadyCapability._new_ifc_global_id()
        resources: List[Dict[str, Any]] = []

        for index, facade in enumerate(contract.entities):
            worksheet = workbook.create_sheet(title=facade.entity_name[:31])
            field_names = [field.name for field in facade.fields]
            worksheet.append(field_names)
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            worksheet.freeze_panes = "A2"

            if index == 0:
                row = dict.fromkeys(field_names, "")
                row.update(
                    {
                        "GlobalId": root_global_id,
                        "Name": instance_name,
                        "CreationDate": datetime.now(timezone.utc)
                        .replace(microsecond=0)
                        .isoformat(),
                        "PredefinedType": "PLANNED",
                    }
                )
                worksheet.append([row.get(name, "") for name in field_names])

            if (
                facade.entity_name == "IfcWorkSchedule"
                and "PredefinedType" in field_names
            ):
                column = field_names.index("PredefinedType") + 1
                validation = DataValidation(
                    type="list",
                    formula1=(
                        '"ACTUAL,BASELINE,PLANNED,USERDEFINED,NOTDEFINED"'
                    ),
                    allow_blank=True,
                )
                worksheet.add_data_validation(validation)
                validation.add(
                    f"{worksheet.cell(row=2, column=column).coordinate}:"
                    f"{worksheet.cell(row=1048576, column=column).coordinate}"
                )

            resources.append(
                self._resource_descriptor(
                    workbook_path, datapackage_path, facade
                )
            )

        workbook.save(workbook_path)
        workbook.close()
        self._write_datapackage(datapackage_path, resources)
        return EntityContextWorkbookArtifact(
            workbook_path=workbook_path,
            datapackage_path=datapackage_path,
            worksheets=tuple(item.entity_name for item in contract.entities),
            root_global_id=root_global_id,
        )

    def _resource_descriptor(
        self,
        workbook_path: Path,
        datapackage_path: Path,
        facade: EntityFacadeContract,
    ) -> Dict[str, Any]:
        return {
            "name": self._snake_case(facade.entity_name),
            "path": Path(
                os.path.relpath(workbook_path, start=datapackage_path.parent)
            ).as_posix(),
            "format": "xlsx",
            "mediatype": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            "dialect": {"excel": {"sheet": facade.entity_name[:31]}},
            "schema": {
                "fields": [
                    {
                        "name": field.name,
                        "type": self._frictionless_type(field.datatype),
                    }
                    for field in facade.fields
                ]
            },
            "entityUri": facade.entity_uri,
            "entityIdentifier": facade.entity_name,
            "facadeUri": facade.facade_uri,
        }

    @staticmethod
    def _write_datapackage(
        datapackage_path: Path,
        resources: List[Dict[str, Any]],
    ) -> None:
        descriptor: Dict[str, Any] = {}
        if datapackage_path.is_file():
            try:
                descriptor = json.loads(
                    datapackage_path.read_text(encoding="utf-8")
                )
            except (json.JSONDecodeError, OSError):
                pass
        descriptor.setdefault("name", "ontobdc_container")
        descriptor["resources"] = resources
        datapackage_path.write_text(
            json.dumps(descriptor, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @staticmethod
    def _frictionless_type(datatype: str) -> str:
        normalized = datatype.lower()
        if normalized in {"boolean", "bool"}:
            return "boolean"
        if normalized in {"integer", "int", "positiveinteger"}:
            return "integer"
        if normalized in {"decimal", "double", "float"}:
            return "number"
        if normalized in {"date", "datetime", "time"}:
            return normalized
        return "string"

    @staticmethod
    def _snake_case(value: str) -> str:
        result = []
        for index, character in enumerate(value):
            if index and character.isupper() and value[index - 1].islower():
                result.append("_")
            result.append(character.lower())
        return "".join(result)
